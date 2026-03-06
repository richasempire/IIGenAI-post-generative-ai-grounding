"""
convert_ice.py — One-time conversion script for the ICE Database Excel file.

Usage:
    cd backend/
    python data/convert_ice.py

Reads : data/ICE_DB.xlsx   (place the file here before running)
Writes: data/ice_materials.json

Sheet   : "ICE Summary"
Col F   : material name  (index 5 in row tuple)
Col G   : CO2e value     (index 6)  — float for material rows, non-numeric for headers
Col J   : comment        (index 9)
Data starts at row 17 (1-indexed / openpyxl convention).
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl not installed.  Run: pip install openpyxl")

# ---------------------------------------------------------------------------
# Paths (all relative to this file so the script works from any cwd)
# ---------------------------------------------------------------------------

DATA_DIR = Path(__file__).resolve().parent
EXCEL_PATH = DATA_DIR / "ICE_DB.xlsx"
JSON_PATH = DATA_DIR / "ice_materials.json"

SOURCE = "ICE Database V4.1 (Oct 2025)"
SHEET_NAME = "ICE Summary"
DATA_START_ROW = 17   # 1-indexed (openpyxl)

# Tuple indices (0-indexed, as returned by values_only iteration)
IDX_F = 5   # material name / category header
IDX_G = 6   # CO₂e value
IDX_J = 9   # comment


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_number(val: object) -> bool:
    """True for int/float values that are NOT booleans."""
    return isinstance(val, (int, float)) and not isinstance(val, bool)


def _text(val: object) -> str:
    return str(val).strip() if val is not None else ""


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert() -> list[dict]:
    if not EXCEL_PATH.exists():
        sys.exit(
            f"Excel file not found at {EXCEL_PATH}\n"
            "Place ICE_DB.xlsx in backend/data/ and re-run."
        )

    print(f"Loading {EXCEL_PATH} …")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True, read_only=True)

    if SHEET_NAME not in wb.sheetnames:
        sys.exit(
            f'Sheet "{SHEET_NAME}" not found.\n'
            f"Available sheets: {wb.sheetnames}"
        )

    ws = wb[SHEET_NAME]
    materials: list[dict] = []
    current_category = "Uncategorised"
    skipped = 0

    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        # Guard against short rows
        if len(row) <= IDX_J:
            row = list(row) + [None] * (IDX_J + 1 - len(row))

        col_f = row[IDX_F]
        col_g = row[IDX_G]
        col_j = row[IDX_J]

        f_text = _text(col_f)
        if not f_text:
            continue  # blank row — skip

        if _is_number(col_g):
            # ── Material row ──────────────────────────────────────────────
            materials.append({
                "category":    current_category,
                "material":    f_text,
                "co2e_per_kg": round(float(col_g), 6),
                "unit":        "kgCO2e/kg",
                "comment":     _text(col_j),
                "source":      SOURCE,
            })
        else:
            # ── Category header ───────────────────────────────────────────
            # Ignore rows that look like footnotes (start with * or †)
            if f_text.startswith(("*", "†", "Note", "Source")):
                skipped += 1
                continue
            current_category = f_text

    wb.close()
    return materials


def main() -> None:
    materials = convert()

    if not materials:
        sys.exit("No materials extracted — check the sheet structure and DATA_START_ROW.")

    with open(JSON_PATH, "w", encoding="utf-8") as fh:
        json.dump(materials, fh, indent=2, ensure_ascii=False)

    # Summary
    categories = {m["category"] for m in materials}
    print(f"✓  Extracted {len(materials)} materials across {len(categories)} categories")
    print(f"✓  Saved → {JSON_PATH}")


if __name__ == "__main__":
    main()
